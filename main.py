import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def generate_hierarchy_decimal_numbering_to_excel(
    initial_path, excel_file="output.xlsx", use_absolute_paths=True
):
    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Directory Structure"

    # Add headers to the first row
    ws.append(["Decimal Number", "File/Directory Name (with Hyperlink)"])

    # Set column widths for better readability
    ws.column_dimensions[get_column_letter(1)].width = 20  # Decimal Number
    ws.column_dimensions[get_column_letter(2)].width = 80  # File/Directory Name

    def get_hyperlink_path(item_path):
        """Returns either an absolute or relative path formatted as a hyperlink."""
        if use_absolute_paths:
            # Convert absolute path to a file URL and encode spaces
            return f"file:///{item_path.replace(' ', '%20')}"
        else:
            # Create a relative path from the initial directory and encode spaces
            relative_path = os.path.relpath(item_path, initial_path)
            #return f"file:///{relative_path.replace(' ', '%20')}"
            return relative_path

    def traverse_directory(path, depth_prefix="1"):
        """Recursively traverse directories and write to Excel with decimal numbering and hyperlinks."""
        items = sorted(os.listdir(path))  # Ensure consistent order

        for index, item in enumerate(items, start=1):
            item_path = os.path.join(path, item)
            # Generate the current decimal key (e.g., "1.1", "1.2.1")
            current_key = f"{depth_prefix}.{index}"

            # Display either just the name or the full path
            display_name = item if not use_absolute_paths else item_path

            # Get the hyperlink path based on the user preference
            link = get_hyperlink_path(item_path)

            # Write the decimal key and hyperlink to the Excel sheet
            cell = ws.cell(row=ws.max_row + 1, column=1, value=current_key)  # Decimal number
            hyperlink_cell = ws.cell(row=cell.row, column=2, value=display_name)  # Name

            # Set the hyperlink and formatting
            hyperlink_cell.hyperlink = link
            hyperlink_cell.font = Font(color="0000FF", underline="single")  # Blue, underlined text

            # If it's a directory, recurse into it
            if os.path.isdir(item_path):
                traverse_directory(item_path, current_key)

    # Start traversal from the initial path
    traverse_directory(initial_path)

    # Save the workbook to the specified Excel file
    wb.save(excel_file)
    print(f"Directory structure with hyperlinks saved to {excel_file}")

# Example usage: Provide the path to the directory you want to scan
initial_path = "C:/Users/innovacion/Proyectos con Ingenieria SAS/Proyectos - General/I000042 Perm. Menores_Entrega N°9"  # Replace with your path
generate_hierarchy_decimal_numbering_to_excel(initial_path, use_absolute_paths=True)
